from types import MethodType
from flask import Flask, request, jsonify
from flask_mysqldb import MySQL

from openpyxl import load_workbook

from openpyxl.styles import Color, PatternFill, Font, Border ,Alignment ,Side
from openpyxl.utils import get_column_letter

import yaml
import json
import csv




app = Flask(__name__)

# db = yaml.load(open('back.yaml'))
# app.config['MYSQL_HOST'] = db['mysql_host'] 
# app.config['MYSQL_USER'] = db['mysql_user']
# app.config['MYSQL_PASSWORD'] = db['mysql_password']
# app.config['MYSQL_DB'] = db['mysql_db']



# mysql = MySQL(app)

# @app.route('/register', methods =['POST'])
# def index():
#     if request.method == 'POST':
#         userDetails = request.form
#         Name = userDetails['Name']
#         Surname = userDetails['Surname']
#         Email = userDetails['Email']
#         Password = userDetails['Password']
#         ConfirmPassword = userDetails['ConfirmPassword']
       
#         data = request.get_json()
        
#         cur = mysql.connection.cursor()
#         cur.execute("INSERT INTO users(Name,Surname,Email,Password,ConfirmPassword) VALUES(%s, %s, %s, %s, %s)", (Name,Surname,Email,Password,ConfirmPassword))
#         mysql.connection.commit() 
#         cur.close()
#         return 'great success'
        

# @app.route("/login", methods =['POST'])
# def login():
#     if request.method == 'POST':
#         userDetails = request.form
#         Email = userDetails['Email']
#         Password = userDetails['Password']
#         data = request.get_json()
        
#         cur = mysql.connection.cursor()
#         count = cur.execute('select * from users where Email=%s and Password=%s', (Email,Password)) 
#         mysql.connection.commit() 
#         cur.close()
#         if count == 0:
#             return "Okay"
#         else:
#             return "User exist"  

@app.route("/items_picked", methods =['POST'])
def items_picked():
    if request.method == 'POST':
        userDetails = request.form
        datas = request.get_json()
        csvfilename=datas[0]["PickListNo"] + "done" + ".csv"
    
        for data in datas[0]["items"] :
            data["ItemNo"] = data.pop("number")
            data["ItemDescription"] = data.pop("description")
            data["Quantity"] = data.pop("quantity")
            data["Location"] = data.pop("location")
            data["KitNo"] = datas[0]["KitNo"]
            data["PickListNo"] = datas[0]["PickListNo"]
            data["Station"] = datas[0]["Station"]
            data["Kitter"] = datas[0]["Kitter"]
            data["UniqueNo"] = datas[0]["UniqueNo"]
        fields = ("KitNo","PickListNo","Station","Kitter","UniqueNo","ItemNo","ItemDescription","Quantity","Location","quantity_picked")

        with open (csvfilename,"w") as file:
            writer = csv.DictWriter(file, fieldnames=fields)
            writer.writeheader()
            for data in datas[0]["items"]:
                writer.writerow(data)
        return json.dumps("done")


@app.route("/list", methods = ['GET'])
def itemlist():
    fields = ("KitNo","PickListNo","Station","Kitter","UniqueNo","ItemNo","ItemDescription","Quantity","Location","SerialNo")
    info = {}
    items = []
    data = []
    csvfilename = request.args.get("csvfilename") + ".csv"
    try:
        with open (csvfilename,"r") as file:
            reader = csv.DictReader( file, fields)
            for row in reader:
                break

            for row in reader:
                info["KitNo"] = row["KitNo"]
                info["PickListNo"] = row["PickListNo"]
                info["Station"] = row["Station"]
                info["Kitter"] = row["Kitter"]
                info["UniqueNo"] = row["UniqueNo"]
                

                

                items.append({"number" : row["ItemNo"] ,
                    "description" : row["ItemDescription"] ,
                    "quantity" : row["Quantity"] ,
                    "location" : row["Location"]})
            
            info["items"]= items
        print(json.dumps([info], indent = 4))
        return json.dumps([info], indent = 4) 
        
    except FileNotFoundError :
        return ('' , 400)          

@app.route("/layout", methods = ['GET'])
def get_layout():
    workbook = load_workbook("practice2.xlsx")
    ws = workbook["Sheet"]

    number_of_segments = ws.cell(row = 1, column = 1).value
    len_racks = ws.cell(row = 2, column = 1).value

    # Program to parse excel file


    def getrow(start_col , start_row  ,  start_with , len ,with_rack):

        
        for i in range(0,len):
            if with_rack : 
                cell1 = ws.cell(row = start_row + 1   , column = start_col + i   )
                value = str(start_with) +"-1"
                values[value] = " ".join(str(cell1.value).split(" ")[0:2])
            start_with +=1
                

    def get_column_racks(start_col , start_row , len ,segment_number , y ):
        for i in range(1,len+1):
            circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
            cell = ws.cell(row = start_row + i - 1 , column = start_col )
            value = str(circle_number) + "-" + str(y)
            values[value] = " ".join(str(cell.value).split(" ")[0:2])
    
    values = {}

    for i in range(0,number_of_segments):   

        getrow(start_col = 4 +i*4 , start_row = 8, start_with = i*3 , len = 3 , with_rack=False)
        get_column_racks(4 + i*4 , 10 , len_racks , i*2 , i*2 + 2)
        get_column_racks(6 + i*4 , 10 , len_racks , i*2 +1 ,  i*2 + 3)
        getrow(start_col = 4 + i*4, start_row = 11 + len_racks  , start_with = i*3 + len_racks*number_of_segments + number_of_segments*3 , len = 3 , with_rack=True)

    data = {}
    data["number_of_segments"] = number_of_segments
    data["number_of_racks"] = len_racks
    data["values"] = values
    for key in sorted(data.items()) :
        print(key)
    return json.dumps(data , indent = 4)

if __name__ == '__main__':
    app.run(host = '192.168.43.146' ,debug = True )    
