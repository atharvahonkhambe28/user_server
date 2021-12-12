from openpyxl import load_workbook

from openpyxl.styles import Color, PatternFill, Font, Border ,Alignment ,Side
from openpyxl.utils import get_column_letter

workbook = load_workbook("practice2.xlsx")
ws = workbook["Sheet"]

number_of_segments = ws.cell(row = 1, column = 1).value
len_racks = ws.cell(row = 2, column = 1).value

# Program to parse excel file


def getrow(start_col , start_row  ,  start_with , len ,with_rack):

    
    for i in range(0,len):
        if with_rack : 
            cell1 = ws.cell(row = start_row + 1   , column = start_col + i   )
            value = str(start_with) +"-1"
            values[value] = cell1.value
        start_with +=1
            

def get_column_racks(start_col , start_row , len ,segment_number , y ):
    for i in range(1,len+1):
        circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
        cell = ws.cell(row = start_row + i - 1 , column = start_col )
        value = str(circle_number) + "-" + str(y)
        values[value] = cell.value

values = {}

for i in range(0,number_of_segments):   

    getrow(start_col = 4 +i*4 , start_row = 8, start_with = i*3 , len = 3 , with_rack=False)
    get_column_racks(4 + i*4 , 10 , len_racks , i*2 , i*2 + 2)
    get_column_racks(6 + i*4 , 10 , len_racks , i*2 +1 ,  i*2 + 3)
    getrow(start_col = 4 + i*4, start_row = 11 + len_racks  , start_with = i*3 + len_racks*number_of_segments + number_of_segments*3 , len = 3 , with_rack=True)

values["number_of_segments"] = number_of_segments
values["number_of_racks"] = len_racks
for key in sorted(values.items()) :
    print(key)