from openpyxl import Workbook

from openpyxl.styles import Color, PatternFill, Font, Border ,Alignment ,Side
from openpyxl.utils import get_column_letter

workbook = Workbook()
ws = workbook.active
# program to create template excel files

number_of_segments = int(input("Enter Number of segments"))
len_racks = int(input("Enter Number of Racks per aisle"))
# number_of_segments = 3
# len_racks = 5
cell = ws.cell(row = 1 , column = 1 , value= number_of_segments)
cell = ws.cell(row = 2 , column = 1 , value= len_racks)

rackFill = PatternFill(fgColor='8EA9DB' , fill_type='solid')
gapFill = PatternFill(fgColor='C9C9C9' , fill_type='solid')
rackAlign = Alignment(horizontal = 'center' , vertical = 'center')
pathFont = Font(color="CDD9EF")
bd = Side(style='thin', color="000000")
rackBorder = Border(left=bd, top=bd, right=bd, bottom=bd)

def name_row(start_col , start_row  ,  start_with , len ,with_rack):

    
    for i in range(0,len):
        cell = ws.cell(row = start_row , column = start_col + i , value= start_with)
        cell.font = pathFont
        cell.alignment = rackAlign
        if with_rack : 
            cell1 = ws.cell(row = start_row + 1   , column = start_col + i ,value = str(start_with) +"-1"  )
            cell1.border = rackBorder
            cell1.alignment = rackAlign
            cell1.fill = rackFill
        start_with +=1
            

def name_column_racks(start_col , start_row , len ,segment_number , y ):
    for i in range(1,len+1):
        circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
        cell = ws.cell(row = start_row + i - 1 , column = start_col , value = str(circle_number) + "-" + str(y))
        cell.fill  = rackFill
        cell.alignment = rackAlign
        cell.border = rackBorder

def add_gap_column(start_col , start_row , len):
    ws.column_dimensions[get_column_letter(start_col)].width = 2
    for i in range(0,len) :
        cell = ws.cell(row = start_row + i, column= start_col)
        cell.fill = gapFill
        


def name_path(start_col , start_row , len ,segment_number):
    for i in range(1,len+1):
        circle_number = (segment_number//2) + i*number_of_segments + number_of_segments*2
        cell = ws.cell(row = start_row + i - 1 , column = start_col , value = circle_number)
        cell.alignment = rackAlign
        cell.font = pathFont



for i in range(0,number_of_segments):   

    name_row(start_col = 4 +i*4 , start_row = 8, start_with = i*3 , len = 3 , with_rack=False)

    name_column_racks(4 + i*4 , 10 , len_racks , i*2 , i*2 + 2)
    name_path(5 + i*4  , 10 , len_racks , i*2 )
    name_column_racks(6 + i*4 , 10 , len_racks , i*2 +1 ,  i*2 + 3)
    add_gap_column(7 + i*4, 8 , len_racks + 5)

    name_row(start_col = 4 + i*4, start_row = 11 + len_racks  , start_with = i*3 + len_racks*number_of_segments + number_of_segments*3 , len = 3 , with_rack=True)


workbook.save("practice4.xlsx")


