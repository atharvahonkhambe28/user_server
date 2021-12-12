import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
workbook = xl.Workbook()
ws = workbook.active
redFill = PatternFill(fgColor='FFFF00' , fill_type='solid')
workSheet.column_dimensions[get_column_letter(4)].width = column_width
ws['A1'].fill = redFill


workbook.save("practice.xlsx")

